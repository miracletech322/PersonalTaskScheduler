from PySide6.QtWidgets import QWidget, QTableWidgetItem, QFileDialog
from PySide6.QtCore import Qt, QFile, QTextStream, QDate

from ui_accountabilityreports import Ui_AccountabilityReports

from openpyxl import Workbook
import global_vars

class AccountabilityReports(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.ui = Ui_AccountabilityReports()
        self.ui.setupUi(self)
        self.parent = parent

        self.ui.btnExport.clicked.connect(self.handleBtnExport)
        self.ui.dateEdit.dateChanged.connect(self.handleDateEdit)

        self.initCSS()
        self.ui.dateEdit.setDate(QDate.currentDate())
        self.loadLogs()
        self.ui.tableWidget.verticalHeader().setVisible(False)
        self.ui.tableWidget.setColumnWidth(1, 140)
        self.ui.tableWidget.setColumnWidth(2, 310)
        self.ui.tableWidget.setColumnWidth(5, 80)
        self.ui.tableWidget.setColumnWidth(6, 80)
        parent.themeChanged.connect(self.initCSS)
    
    def initCSS(self):
        url = ":/Resources/accountabilityreports.qss"
        if global_vars.app_theme == "Light Mode":
            url = ":/Resources/accountabilityreports_light.qss"

        file = QFile(url)
        if file.open(QFile.ReadOnly | QFile.Text):
            stream = QTextStream(file)
            stylesheet = stream.readAll()
            self.setStyleSheet(stylesheet)
            file.close()
    
    def loadLogs(self):
        date = self.ui.dateEdit.date().toPython()
        logs = self.parent.funcLoadLogs(date)
        self.ui.tableWidget.setRowCount(0)
        row = 0
        for log in logs:
            self.ui.tableWidget.insertRow(row)
            item = QTableWidgetItem(log['task']['time'])
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 0, item)

            item = QTableWidgetItem(log['task']['title'])
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 1, item)

            item = QTableWidgetItem(log['task']['description'])
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 2, item)

            user_str = str(log["user"]['username']) if log.get("user") else "None"
            item = QTableWidgetItem(user_str)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 3, item)

            hours = ""
            if log['status'] == "Yes" and log['task']['description'] == "BREAKFAST/MEDS COMPLETED?":
                hours = "1hr(s)"
            if log['status'] == "No" and log['task']['description'] == "BREAKFAST/MEDS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "MEDS/PRESSURE/WATER COMPLETED?":
                hours = "0.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "MEDS/PRESSURE/WATER COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "INTERACTION/ACTIVITIES/LUNCH PREP/MEDS COMPLETED?":
                hours = "2hr(s)"
            if log['status'] == "No" and log['task']['description'] == "INTERACTION/ACTIVITIES/LUNCH PREP/MEDS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "LUNCH/KITCHEN/MEDS COMPLETED?":
                hours = "1hr(s)"
            if log['status'] == "No" and log['task']['description'] == "LUNCH/KITCHEN/MEDS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "SNACKS/COOK DINNER COMPLETED?":
                hours = "3.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "SNACKS/COOK DINNER COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "DINNER/MEDS/BREAK/KITCHEN/DISHES COMPLETED?":
                hours = "1.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "DINNER/MEDS/BREAK/KITCHEN/DISHES COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "PUT TO BED/MEDS/CHARTING/CHECK RESIDENTS COMPLETED?":
                hours = "1hr(s)"
            if log['status'] == "No" and log['task']['description'] == "PUT TO BED/MEDS/CHARTING/CHECK RESIDENTS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "BREAK/ALL OTHER ASSIGNMENTS/MEDS COMPLETED?":
                hours = "1.25hr(s)"
            if log['status'] == "No" and log['task']['description'] == "BREAK/ALL OTHER ASSIGNMENTS/MEDS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "LAUNDRY STARTED?":
                hours = "0.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "LAUNDRY STARTED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "RESIDENTS CHECKED?":
                hours = "0.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "RESIDENTS CHECKED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "LAUNDRY COMPLETED?":
                hours = "1.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "LAUNDRY COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "RESIDENTS CHECKED?":
                hours = "0.5hr(s)"
            if log['status'] == "No" and log['task']['description'] == "RESIDENTS CHECKED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "BATHROOMS COMPLETED?":
                hours = "1hr(s)"
            if log['status'] == "No" and log['task']['description'] == "BATHROOMS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "ROOMS SWEPT/FLOORS MOPPED COMPLETED?":
                hours = "1hr(s)"
            if log['status'] == "No" and log['task']['description'] == "ROOMS SWEPT/FLOORS MOPPED COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "BREAK/RESIDENTS CHECKED?":
                hours = "0.25hr(s)"
            if log['status'] == "No" and log['task']['description'] == "BREAK/RESIDENTS CHECKED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "RESIDENTS CHECKED/PANTRY/CLEANING/TRASH/ROOMS COMPLETED?":
                hours = "1.25hr(s)"
            if log['status'] == "No" and log['task']['description'] == "RESIDENTS CHECKED/PANTRY/CLEANING/TRASH/ROOMS COMPLETED?":
                hours = "0hr(s)"

            if log['status'] == "Yes" and log['task']['description'] == "RESIDENTS CLEANED/BATHED/BRUSHED/READY/BEDS COMPLETED?":
                hours = "1.25hr(s)"
            if log['status'] == "No" and log['task']['description'] == "RESIDENTS CLEANED/BATHED/BRUSHED/READY/BEDS COMPLETED?":
                hours = "0hr(s)"

            item = QTableWidgetItem(log['status'])
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 4, item)

            time_str = log["timestamp"].strftime("%H:%M")
            item = QTableWidgetItem(time_str)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 5, item)

            item = QTableWidgetItem(hours)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.ui.tableWidget.setItem(row, 6, item)

            row = row + 1

    def handleBtnExport(self):
        str = self.ui.dateEdit.date().toPython().strftime("%Y-%m-%d")

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            str,
            "Excel Files (*.xlsx);"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Alarm Time"

        ws["A1"] = "Alarm Time"
        ws["B1"] = "Title"
        ws["C1"] = "Description"
        ws["D1"] = "User"
        ws["E1"] = "Processed"
        ws["F1"] = "Clicked"
        ws["G1"] = "Hrs Worked"
        row = self.ui.tableWidget.rowCount()
        for i in range(row):
            ws.append((
                self.ui.tableWidget.item(i, 0).text(),
                self.ui.tableWidget.item(i, 1).text(),
                self.ui.tableWidget.item(i, 2).text(),
                self.ui.tableWidget.item(i, 3).text(),
                self.ui.tableWidget.item(i, 4).text(),
                self.ui.tableWidget.item(i, 5).text(),
                self.ui.tableWidget.item(i, 6).text(),
            ))
        wb.save(file_path)

    def handleDateEdit(self, date):
        self.loadLogs()
